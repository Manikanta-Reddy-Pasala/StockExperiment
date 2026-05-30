import sys; sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parents[2]))
import openpyxl,re,json
wb=openpyxl.load_workbook('/Users/manip/Downloads/Nifty500_Rebalancing_v2.xlsx',read_only=True,data_only=True)
# name->symbol map from all per-year Constituents sheets
name2sym={}
for y in range(2021,2027):
    ws=wb[f'{y} Constituents']; rows=list(ws.iter_rows(values_only=True))
    hi=next(i for i,r in enumerate(rows) if r and 'Symbol' in [str(c) for c in r if c])
    cj=[str(c) for c in rows[hi]]; ci=cj.index('Company Name'); si=cj.index('Symbol')
    for r in rows[hi+1:]:
        if r and r[si] and r[ci]:
            name2sym[str(r[ci]).strip().lower().rstrip('.')]=str(r[si]).strip()
def syms_from_delta(cell):
    if not cell: return set()
    out=set()
    for line in str(cell).split('\n'):
        line=line.strip().lstrip('+-').strip()
        if not line or 'Reference snapshot' in line: continue
        m=re.search(r'\(([A-Z0-9&\-]{2,15})\)\s*$', line)  # (SYMBOL) at end
        if m: out.add(m.group(1)); continue
        nm=re.sub(r'\s*\(.*?\)\s*$','',line).strip().lower().rstrip('.')
        if nm in name2sym: out.add(name2sym[nm])
        else:
            # try without Ltd etc
            nm2=re.sub(r'\s+ltd\.?$','',nm).strip()
            for k,v in name2sym.items():
                if k.startswith(nm2[:20]) and len(nm2)>8: out.add(v); break
    return out
# anchor = Year-wise Constituents matrix (treat as that year's primary state)
ws=wb['Year-wise Constituents']; rows=list(ws.iter_rows(values_only=True))
hi=next(i for i,r in enumerate(rows) if r and 'Symbol' in [str(c) for c in r])
hdr=rows[hi]; sj=hdr.index('Symbol'); yc={str(c).strip():j for j,c in enumerate(hdr) if c and str(c).strip() in ('2021','2022','2023','2024','2025','2026')}
anchor={}
for y,j in yc.items():
    anchor[y]=set(str(r[sj]).strip() for r in rows[hi+1:] if r and r[sj] and str(r[j]).strip() not in('—','-','None',''))
# Summary deltas
ws=wb['Year-wise Summary']; srows=list(ws.iter_rows(values_only=True))
deltas={}  # (year,review)->(adds,removes)
for r in srows:
    if r and r[0] and str(r[0]).strip() in ('2021','2022','2023','2024','2025','2026') and r[1]:
        y=str(r[0]).strip(); rev=str(r[1]).strip()  # 'Mar 2021'
        deltas[(y,rev)]=(syms_from_delta(r[3]),syms_from_delta(r[4]))
# Build Mar+Sep: anchor = the matrix year (assume ~Mar/primary). Sep = anchor + sepAdds - sepRemoves
snaps={}
for y in ('2021','2022','2023','2024','2025','2026'):
    base=anchor[y]
    snaps[f"{y}0331"]=set(base)
    sa,sr=deltas.get((y,f"Sep {y}"),(set(),set()))
    snaps[f"{y}0930"]=(base|sa)-sr if (sa or sr) else set(base)
for d in sorted(snaps):
    print(f"  N500 {d[:4]}-{d[4:6]}: {len(snaps[d])}")
json.dump({k:sorted(v) for k,v in snaps.items()},open('/tmp/n500_semi.json','w'))
print("name2sym entries:",len(name2sym))
